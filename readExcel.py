import time
from openpyxl import  Workbook,load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By


driver = webdriver.Chrome(executable_path='C:\webdriver\chromedriver.exe')
# driver = webdriver.Chrome()
driver.get("https://www.saucedemo.com/")
book=load_workbook('DataFile.xlsx')
sheet=book.active
rows=sheet.max_row
cols=sheet.max_column
username_element=driver.find_element(By.ID,'user-name')
password_element=driver.find_element(By.ID,'password')

for i in range(2,rows):
    # print(str(sheet.cell(i,1.value))
    # print(sheet.cell(i,1.value))
    username=str(sheet.cell(i,1).value)
    username_element.clear()
    username_element.send_keys(username)
    time.sleep(5)
    password=sheet.cell(i,2).value
    password_element.clear()
    password_element.send_keys(password)
    time.sleep(5)
    # print(password)


# driver.find_element(By.ID,'user-name').send_keys(username)
# driver.find_element(By.ID,'password').send_keys(password)
# driver.find_element(By.ID,'login-button').click()

# time.sleep(10)